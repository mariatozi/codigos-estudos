
import openpyxl

# criar uma planilha
book = openpyxl.Workbook()
# visualizar páginas existentes na planilha
print(book.sheetnames)
# criar uma página
book.create_sheet('Meus Computadores')
# selecionar uma página
pc_pages = book['Meus Computadores']
# adicionando o 'cabeçalho' dos dados inseridos na planilha
pc_pages.append(['Eletrônica', 'memória RAM', 'preço'])
# adicionar dados na página (adicionar em linhas) 
pc_pages.append(['Computador 1', '8gb RAM', 'R$2500'])
pc_pages.append(['Computador 2', '16gb RAM', 'R$5500'])
pc_pages.append(['Computador 3', '32gb RAM', 'R$8500'])
# salvar a planilha e quando for realizando as alterações
book.save('Meus Computadores.xlsx')

import openpyxl

# carregando arquivo
book = openpyxl.load_workbook('Meus Computadores.xlsx')
# selecionando uma página
pc_pages = book['Meus Computadores']
# imprimindo os dados de cada linha
for rows in pc_pages.iter_rows(min_row=2, max_row=5):
    for cell in rows:
        print(cell.value )

# salvar a planilha e as alterações
book.save('Meus Computadores.xlsx')
